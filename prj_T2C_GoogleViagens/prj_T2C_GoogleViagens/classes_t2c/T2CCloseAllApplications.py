from botcity.web import WebBot, Browser
from botcity.core import DesktopBot
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CMaestro import T2CMaestro, LogLevel, ErrorType
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CExceptions import BusinessRuleException
from openpyxl import load_workbook
import pandas as pd

class T2CCloseAllApplications:
    """
    Classe para fechar todos os aplicativos no final da automação.

    Parâmetros:
    
    Retorna:
    """
    def __init__(self, arg_dictConfig:dict, arg_clssMaestro:T2CMaestro, arg_botWebbot:WebBot=None, arg_botDesktopbot:DesktopBot=None):
        """
        Inicializa a classe.

        Parâmetros:
        - arg_dictConfig (dict): dicionário de configuração.
        - arg_clssMaestro (T2CMaestro): instância da classe T2CMaestro.
        - arg_botWebbot (WebBot): instância do bot WebBot (opcional, default=None).
        - arg_botDesktopbot (DesktopBot): instância do bot DesktopBot (opcional, default=None).

        Retorna:
        """
        if(arg_botWebbot is None and arg_botDesktopbot is None): raise Exception("Não foi possível inicializar a classe, forneça pelo menos um bot")
        else:
            self.var_botWebbot = arg_botWebbot
            self.var_botDesktopbot = arg_botDesktopbot
            self.var_dictConfig = arg_dictConfig
            self.var_clssMaestro = arg_clssMaestro

    def execute(self):
        """
        Executa o fechamento de todos os aplicativos necessários, apenas com a estrutura em código.

        Observação:
        - Edite o valor da variável `var_intMaxTentativas` no arquivo Config.xlsx.

        Parâmetros:
        
        Retorna:

        Raises:
        - BusinessRuleException: em caso de erro de regra de negócio.
        - Exception: em caso de erro geral.
        """

        #Edite o valor dessa variável a no arquivo Config.xlsx
        var_intMaxTentativas = self.var_dictConfig["MaxRetryNumber"]

        for var_intTentativa in range(var_intMaxTentativas):
            try:
                self.var_clssMaestro.write_log("Finalizando todos os processos, tentativa " + (var_intTentativa+1).__str__())
                #Insira aqui seu código para fechar os aplicativos

                caminho_excel = 'prj_T2C_GoogleViagens/Precos_viagem.xlsx'
                var_wbkSintetico = load_workbook(caminho_excel)
                
                df = pd.read_excel(caminho_excel, sheet_name='Todos')
        
                # Garantir que a coluna 'Preço' é interpretada corretamente (remover 'R$' e converter para numérico)
                df['Preço'] = df['Preço'].replace('[R$]', '', regex=True).astype(float)
                
                # Ordenar por preço e pegar as 10 linhas com os preços mais baixos
                df_sorted = df.sort_values(by='Preço').head(10)
                
                # Carregar o arquivo Excel para escrita
                with pd.ExcelWriter(caminho_excel, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Carregar a planilha existente
                    var_wbkSintetico = writer.book
                    if 'Baratos' not in var_wbkSintetico.sheetnames:
                        var_wshtBaratos = var_wbkSintetico.create_sheet(title='Baratos')
                        # Adiciona os cabeçalhos
                        var_wshtBaratos.append(['País', 'Cidade', 'Preço'])
                    else:
                        var_wshtBaratos = var_wbkSintetico['Baratos']
                        # Encontra a primeira linha vazia após os cabeçalhos
                        start_row = var_wshtBaratos.max_row + 1

                    # Adiciona os dados na planilha 'Baratos'
                    for row in df_sorted.itertuples(index=False, name=None):
                        var_wshtBaratos.append(row)

                    # Salvar o arquivo
                    var_wbkSintetico.save(caminho_excel)
                    self.var_botWebbot.close_page()
                    break

            except BusinessRuleException as exception:
                self.var_clssMaestro.write_log(arg_strMensagemLog="Erro de negócio: " + str(exception), arg_enumLogLevel=LogLevel.ERROR, arg_enumErrorType=ErrorType.BUSINESS_ERROR)
 
                raise
            except Exception as exception:
                self.var_clssMaestro.write_log(arg_strMensagemLog="Erro, tentativa " + (var_intTentativa+1).__str__() + ": " + str(exception), arg_enumLogLevel=LogLevel.ERROR, arg_enumErrorType=ErrorType.APP_ERROR)
                
                if(var_intTentativa+1 == var_intMaxTentativas): raise
                else: 
                    #Incluir aqui seu código para tentar novamente
                    
                    continue
            else:
                self.var_clssMaestro.write_log("Aplicativos finalizados, continuando processamento...")
                break
            