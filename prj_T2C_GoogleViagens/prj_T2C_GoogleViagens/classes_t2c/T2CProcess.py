from botcity.web import WebBot, Browser
from botcity.core import DesktopBot
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CMaestro import T2CMaestro, LogLevel, ErrorType
from prj_T2C_GoogleViagens.classes_t2c.utils.T2CExceptions import BusinessRuleException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import pandas as pd 
from openpyxl import load_workbook

class T2CProcess:
    def __init__(self, arg_dictConfig:dict, arg_clssMaestro:T2CMaestro, arg_botWebbot:WebBot=None, arg_botDesktopbot:DesktopBot=None):
        if(arg_botWebbot is None and arg_botDesktopbot is None): 
            raise Exception("Não foi possível inicializar a classe, forneça pelo menos um bot")
        else:
            self.var_botWebbot = arg_botWebbot
            self.var_botDesktopbot = arg_botDesktopbot
            self.var_dictConfig = arg_dictConfig
            self.var_clssMaestro = arg_clssMaestro

    def preencher_dados(self, arg_tplQueueItem):
        # De onde
        pesquisa = self.var_botWebbot.find_element('//input[@aria-label="De onde?"]', By.XPATH)
        pesquisa.clear()
        pesquisa.send_keys('São Paulo')
        time.sleep(2)
        
        resultado_pesquisa = self.var_botWebbot.find_element('/html/body/c-wiz[2]/div/div[2]/div/c-wiz/div[2]/div/div/div[1]/div[1]/section/div/div[1]/div[1]/div[1]/div[2]/div[1]/div/div[1]/div/div[2]/div[3]/ul/li[1]', By.XPATH)
        resultado_pesquisa.click()

        # para onde
        pesquisa = self.var_botWebbot.find_element('//input[@aria-label="Para onde?"]', By.XPATH)
        pesquisa.send_keys(arg_tplQueueItem[1])
        time.sleep(2)

        # Escolher o país com o ícone de globo
        pais_com_globo = self.var_botWebbot.find_element(f'//li[@aria-label="{arg_tplQueueItem[1]}"]', By.XPATH)
        pais_com_globo.click()

        #Click no campo de data
        input_data = self.var_botWebbot.find_element('//div[@title="Viagem de uma semana nos próximos seis meses"]', By.XPATH)
        input_data.click()
        time.sleep(2)
        
        #Click na aba Datas específicas
        aba_data_especifica = self.var_botWebbot.find_element('//*[@id="sNlbpb"]', By.XPATH)
        aba_data_especifica.click()
        time.sleep(2)
        
        #Click no input data Início
        data_inicio_input = self.var_botWebbot.find_element('//*[@id="ow78"]/div[2]/div/div[2]/div/div[2]/span/div/div[1]/div/div[1]/div/input', By.XPATH)
        data_inicio_input.send_keys('20/08/2024')
        data_inicio_input.send_keys(Keys.ENTER)
        time.sleep(2)
        
        #Click no input data Fim
        data_fim_input = self.var_botWebbot.find_element('//*[@id="ow78"]/div[2]/div/div[2]/div/div[2]/span/div/div[1]/div/div[2]/div/input', By.XPATH)
        data_fim_input.send_keys('23/08/2024')
        data_fim_input.send_keys(Keys.ENTER)
        time.sleep(2)
        
        #Click no botão Confirmar
        botao_concluido = self.var_botWebbot.find_element('(//span[text()="Concluído"])[2]', By.XPATH)
        botao_concluido.click()   

    def extrair_informacoes(self, pais):
        index = 1
        info_cidades = []
        
        # Conjunto para armazenar nomes de cidades já vistas
        cidades_vistas = set()
        
        while True:               
            xpath_nome_cidade = f'/html/body/c-wiz[2]/div/div[2]/div/c-wiz/div[2]/div/div/div[1]/main/div/div[2]/div/ol/li[{index}]/div/div[2]/div[1]/h3'
            xpath_preco_cidade = f'//*[@id="yDmH0d"]/c-wiz[2]/div/div[2]/div/c-wiz/div[2]/div/div/div[1]/main/div/div[2]/div/ol/li[{index}]/div/div[2]/div[2]/div[1]/div[1]/span'
            
            try:                
                nome_cidade = self.var_botWebbot.find_element(xpath_nome_cidade, By.XPATH).text
                preco_cidade = self.var_botWebbot.find_element(xpath_preco_cidade, By.XPATH).text
            
                # Adicionar o nome da cidade ao conjunto e à lista
                cidades_vistas.add(nome_cidade)

                # Adicionar os dados à lista
                cidade_preço = {
                    'País': pais,
                    'Cidade': nome_cidade,
                    'Preço': preco_cidade
                }
                
                info_cidades.append(cidade_preço)
                print("Dados coletados:", cidade_preço)
                index += 1
                time.sleep(2)
            except Exception as e:
                print("Elemento não encontrado. Encerrando o loop.")
                break
        
        # Salvar os dados em um arquivo Excel
        self.salvar_em_excel(info_cidades)        
        
    def salvar_em_excel(self, dados):
        caminho_excel_passagens = 'prj_T2C_GoogleViagens/Precos_viagem.xlsx'

        var_wbkSintetico = load_workbook(caminho_excel_passagens)
        var_wshtSintetico = var_wbkSintetico.active
        
        # Seleciona a planilha 'Todos' ou cria uma nova se não existir
        if 'Todos' in var_wbkSintetico.sheetnames:
            var_wshtSintetico = var_wbkSintetico['Todos']
        else:
            var_wshtSintetico = var_wbkSintetico.create_sheet(title='Todos')

        # Encontrando a próxima linha vazia
        var_intIndexNewline = var_wshtSintetico.max_row + 1

        # Adicionando os dados
        for dado in dados:
            var_wshtSintetico.append([
                dado.get('País', ''),
                dado.get('Cidade', ''),
                dado.get('Preço', '')
            ])

        # Salvando e fechando o arquivo
        var_wbkSintetico.save(caminho_excel_passagens)
        var_wbkSintetico.close()
        
    def execute(self, arg_tplQueueItem:tuple):
        # Preencher dados
        self.preencher_dados(arg_tplQueueItem)
             
        #Chamada para classe de extrair dados da cidade
        self.extrair_informacoes(arg_tplQueueItem[1])